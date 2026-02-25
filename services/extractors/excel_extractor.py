import csv
import io

import openpyxl


async def extract_excel(file_path: str) -> str:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    content = "--- SPREADSHEET CONTENT ---\n"

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        content += f"\n[Sheet: {sheet_name}]\n"
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            values = [str(cell) if cell is not None else "" for cell in row]
            if any(v for v in values):
                content += f"Row {row_idx}: {' | '.join(values)}\n"

    wb.close()
    return content


async def extract_csv(file_path: str) -> str:
    content = "--- CSV CONTENT ---\n"
    # Try utf-8-sig first (handles BOM), fallback to latin-1
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            with open(file_path, "r", encoding=encoding) as f:
                reader = csv.reader(f)
                for row_idx, row in enumerate(reader, 1):
                    content += f"Row {row_idx}: {' | '.join(row)}\n"
            return content
        except UnicodeDecodeError:
            continue
    return content
